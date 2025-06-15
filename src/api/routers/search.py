"""
Search API endpoints

Provides comprehensive search functionality with intent detection,
filtering, autocomplete, and analytics integration.
"""

import asyncio
import time
from typing import List, Dict, Any
from fastapi import APIRouter, Depends, HTTPException, Query, BackgroundTasks
from fastapi.responses import StreamingResponse
import structlog

from ..models import *
from ..main import get_search_engine, get_current_user, get_metrics
from ...search.unified_search import UnifiedSearchEngine

logger = structlog.get_logger(__name__)

router = APIRouter()

@router.post("/query", response_model=SearchResponse)
async def search_knowledge(
    request: SearchRequest,
    background_tasks: BackgroundTasks,
    search_engine: UnifiedSearchEngine = Depends(get_search_engine),
    user = Depends(get_current_user),
    metrics = Depends(get_metrics)
):
    """
    Perform intelligent search across all indexed content
    
    This endpoint provides the main search functionality with:
    - Automatic intent detection
    - Semantic and exact search
    - Result ranking and filtering
    - Query suggestions
    """
    start_time = time.time()
    
    try:
        logger.info(
            "Search request received",
            user_id=user.id,
            query=request.query,
            intent=request.intent,
            max_results=request.max_results
        )
        
        # Perform search
        search_result = await search_engine.search(
            query=request.query,
            intent=request.intent.value if request.intent else None,
            filters=request.filters or {},
            max_results=request.max_results,
            min_score=request.min_score,
            user_id=user.id
        )
        
        # Convert to API response format
        results = []
        for item in search_result.get('results', []):
            results.append(SearchResultItem(
                id=item['id'],
                title=item.get('title', ''),
                content=item.get('content', ''),
                score=item.get('score', 0.0),
                book_id=item.get('book_id', ''),
                book_title=item.get('book_title', ''),
                page_number=item.get('page_number'),
                chunk_type=item.get('chunk_type', 'text'),
                metadata=item.get('metadata', {}),
                highlights=item.get('highlights', [])
            ))
        
        processing_time = (time.time() - start_time) * 1000
        
        response = SearchResponse(
            query=request.query,
            intent=request.intent.value if request.intent else search_result.get('detected_intent'),
            results=results,
            total_found=search_result.get('total_found', len(results)),
            processing_time_ms=processing_time,
            suggestions=search_result.get('suggestions', []),
            filters_applied=search_result.get('filters_applied', {})
        )
        
        # Log analytics in background
        background_tasks.add_task(
            log_search_analytics,
            metrics,
            user.id,
            request.query,
            len(results),
            processing_time,
            request.intent
        )
        
        logger.info(
            "Search completed",
            user_id=user.id,
            query=request.query,
            results_count=len(results),
            processing_time_ms=processing_time
        )
        
        return response
        
    except Exception as e:
        logger.error(
            "Search failed",
            user_id=user.id,
            query=request.query,
            error=str(e),
            exc_info=True
        )
        raise HTTPException(status_code=500, detail=f"Search failed: {str(e)}")

@router.get("/autocomplete", response_model=AutocompleteResponse)
async def autocomplete_query(
    q: str = Query(..., min_length=1, max_length=100, description="Partial query"),
    max_suggestions: int = Query(5, ge=1, le=20, description="Maximum suggestions"),
    search_engine: UnifiedSearchEngine = Depends(get_search_engine),
    user = Depends(get_current_user)
):
    """
    Get autocomplete suggestions for partial queries
    
    Returns intelligent suggestions based on:
    - Query history
    - Popular searches
    - Trading terminology
    - Content analysis
    """
    try:
        suggestions = await search_engine.get_suggestions(
            partial_query=q,
            max_suggestions=max_suggestions,
            user_id=user.id
        )
        
        return AutocompleteResponse(
            query=q,
            suggestions=suggestions
        )
        
    except Exception as e:
        logger.error("Autocomplete failed", query=q, error=str(e))
        raise HTTPException(status_code=500, detail=f"Autocomplete failed: {str(e)}")

@router.get("/context/{chunk_id}")
async def get_chunk_context(
    chunk_id: str,
    context_size: int = Query(2, ge=1, le=5, description="Number of chunks before/after"),
    search_engine: UnifiedSearchEngine = Depends(get_search_engine),
    user = Depends(get_current_user)
):
    """
    Get surrounding context for a specific chunk
    
    Returns chunks before and after the specified chunk to provide
    better understanding of the content context.
    """
    try:
        context_result = await search_engine.get_chunk_context(
            chunk_id=chunk_id,
            context_size=context_size,
            user_id=user.id
        )
        
        if not context_result:
            raise HTTPException(status_code=404, detail="Chunk not found")
        
        return {
            "chunk_id": chunk_id,
            "context": context_result,
            "total_chunks": len(context_result)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Chunk context retrieval failed", chunk_id=chunk_id, error=str(e))
        raise HTTPException(status_code=500, detail=f"Context retrieval failed: {str(e)}")

@router.get("/similar/{result_id}")
async def find_similar_content(
    result_id: str,
    max_results: int = Query(10, ge=1, le=50),
    search_engine: UnifiedSearchEngine = Depends(get_search_engine),
    user = Depends(get_current_user)
):
    """
    Find content similar to a specific search result
    
    Useful for discovering related concepts, implementations,
    or alternative approaches to trading strategies.
    """
    try:
        similar_results = await search_engine.find_similar(
            result_id=result_id,
            max_results=max_results,
            user_id=user.id
        )
        
        results = []
        for item in similar_results:
            results.append(SearchResultItem(
                id=item['id'],
                title=item.get('title', ''),
                content=item.get('content', ''),
                score=item.get('score', 0.0),
                book_id=item.get('book_id', ''),
                book_title=item.get('book_title', ''),
                page_number=item.get('page_number'),
                chunk_type=item.get('chunk_type', 'text'),
                metadata=item.get('metadata', {}),
                highlights=item.get('highlights', [])
            ))
        
        return SearchResponse(
            query=f"Similar to: {result_id}",
            results=results,
            total_found=len(results),
            processing_time_ms=0.0  # Not tracked for similarity
        )
        
    except Exception as e:
        logger.error("Similar content search failed", result_id=result_id, error=str(e))
        raise HTTPException(status_code=500, detail=f"Similar search failed: {str(e)}")

@router.post("/feedback")
async def submit_search_feedback(
    result_id: str,
    rating: int = Query(..., ge=1, le=5, description="Rating 1-5"),
    query: str = Query(..., description="Original search query"),
    feedback: str = Query(None, description="Optional feedback text"),
    search_engine: UnifiedSearchEngine = Depends(get_search_engine),
    user = Depends(get_current_user)
):
    """
    Submit feedback on search result relevance
    
    This helps improve the ranking algorithm through
    learning to rank techniques.
    """
    try:
        await search_engine.submit_feedback(
            user_id=user.id,
            query=query,
            result_id=result_id,
            rating=rating,
            feedback=feedback
        )
        
        logger.info(
            "Search feedback received",
            user_id=user.id,
            result_id=result_id,
            rating=rating,
            query=query
        )
        
        return {"success": True, "message": "Feedback submitted successfully"}
        
    except Exception as e:
        logger.error("Feedback submission failed", error=str(e))
        raise HTTPException(status_code=500, detail=f"Feedback submission failed: {str(e)}")

@router.get("/trending")
async def get_trending_queries(
    period: str = Query("24h", pattern="^(1h|24h|7d|30d)$"),
    limit: int = Query(20, ge=1, le=100),
    search_engine: UnifiedSearchEngine = Depends(get_search_engine),
    user = Depends(get_current_user)
):
    """
    Get trending search queries
    
    Shows what other users are searching for, helping
    discover popular topics and strategies.
    """
    try:
        trending = await search_engine.get_trending_queries(
            period=period,
            limit=limit
        )
        
        return {
            "period": period,
            "trending_queries": trending,
            "generated_at": datetime.utcnow()
        }
        
    except Exception as e:
        logger.error("Trending queries failed", error=str(e))
        raise HTTPException(status_code=500, detail=f"Trending queries failed: {str(e)}")

@router.get("/export/{format}")
async def export_search_results(
    format: str,
    query: str = Query(..., description="Search query to export"),
    max_results: int = Query(1000, ge=1, le=10000),
    search_engine: UnifiedSearchEngine = Depends(get_search_engine),
    user = Depends(get_current_user)
):
    """
    Export search results in various formats
    
    Useful for analysis, reporting, or integration with
    other trading systems.
    """
    try:
        # Perform search
        search_result = await search_engine.search(
            query=query,
            max_results=max_results,
            user_id=user.id
        )
        
        # Generate export
        if format == "csv":
            content = await generate_csv_export(search_result['results'])
            media_type = "text/csv"
            filename = f"search_results_{int(time.time())}.csv"
        elif format == "json":
            content = await generate_json_export(search_result['results'])
            media_type = "application/json"
            filename = f"search_results_{int(time.time())}.json"
        elif format == "xlsx":
            content = await generate_xlsx_export(search_result['results'])
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"search_results_{int(time.time())}.xlsx"
        
        return StreamingResponse(
            iter([content]),
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error("Export failed", format=format, query=query, error=str(e))
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

# Helper functions
async def log_search_analytics(metrics, user_id: str, query: str, 
                             results_count: int, processing_time: float, 
                             intent: Optional[SearchIntent]):
    """Log search analytics in background"""
    try:
        await metrics.log_search(
            user_id=user_id,
            query=query,
            results_count=results_count,
            processing_time=processing_time,
            intent=intent.value if intent else None
        )
    except Exception as e:
        logger.error("Failed to log search analytics", error=str(e))

async def generate_csv_export(results: List[Dict[str, Any]]) -> str:
    """Generate CSV export of search results"""
    import csv
    import io
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=[
        'id', 'title', 'content', 'score', 'book_title', 'page_number', 'chunk_type'
    ])
    
    writer.writeheader()
    for result in results:
        writer.writerow({
            'id': result['id'],
            'title': result.get('title', ''),
            'content': result.get('content', '')[:500],  # Truncate for CSV
            'score': result.get('score', 0.0),
            'book_title': result.get('book_title', ''),
            'page_number': result.get('page_number', ''),
            'chunk_type': result.get('chunk_type', 'text')
        })
    
    return output.getvalue()

async def generate_json_export(results: List[Dict[str, Any]]) -> str:
    """Generate JSON export of search results"""
    import json
    return json.dumps(results, indent=2, default=str)

async def generate_xlsx_export(results: List[Dict[str, Any]]) -> bytes:
    """Generate Excel export of search results"""
    import openpyxl
    from io import BytesIO
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Search Results"
    
    # Headers
    headers = ['ID', 'Title', 'Content', 'Score', 'Book Title', 'Page', 'Type']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    for row, result in enumerate(results, 2):
        ws.cell(row=row, column=1, value=result['id'])
        ws.cell(row=row, column=2, value=result.get('title', ''))
        ws.cell(row=row, column=3, value=result.get('content', '')[:1000])
        ws.cell(row=row, column=4, value=result.get('score', 0.0))
        ws.cell(row=row, column=5, value=result.get('book_title', ''))
        ws.cell(row=row, column=6, value=result.get('page_number', ''))
        ws.cell(row=row, column=7, value=result.get('chunk_type', 'text'))
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()